registrar = {

}

from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt

file_path = "materials.xlsx"
workbook = load_workbook(file_path)
file = pd.ExcelFile(file_path)
workbook = load_workbook(file_path)

def molarmass():
    grams = float(input("How many grams are there to add? "))
    mol = float(input("How many moles are there to add?"))
    molar = grams/mol
    print(f"{molar:.4f} is the molar mass")
    return grams, mol, molar

def firstcreation():
    name = input("New Project name: ")
    workbook.create_sheet(title=name)
    workbook.save(file_path)
    additions = int(input("Would you like to add to your new project (0), or close the program, (1)? "))
    global file
    file = pd.ExcelFile(file_path)
    if additions == 0:
        addition()

def addition():
    project = input("Which project (sheet) are you adding this to? ")
    material = input("What material would you like to add? ")
    compound = input("What is the compound's chemical formula? ")

    while True:
        grams, mol, molar = molarmass()
        wrong = int(input("correct? (0 for yes 1 for no) "))
        if wrong == 0:
            break
        else:
            print("Re-entering values")

    project = input("Which project (sheet) are you adding this to? ")

    sheet = workbook[project]
    sheet.append([material, compound, grams, mol, molar])
    workbook.save(file_path)

    row = [material, compound, grams, mol, molar]


    registrar[project].append(row)

    print(f"Added {material} to project '{project}'")

def plotmass(project):
    df = pd.read_excel(file_path, sheet_name=project)

    labels = df.iloc[:, 0]
    masses = df.iloc[:, 2]

    plt.figure(figsize=(6, 6))
    plt.pie(masses, labels=labels, autopct='%1.1f%%', startangle=90)
    plt.title(f"Mass Breakdown for Project: {project}")
    plt.tight_layout()
    plt.show()


###########################################################################################################################################


print(file.sheet_names)

while True:
    create = int(input("Is your project here (0) or would you like to create a new project (1)?"))

    if create == 0:        
        for i in file.sheet_names:
            registrar[i] = pd.read_excel(file, sheet_name=i)
            print(registrar[i])
        index = int(input("Which project would you like to view (1, 2, 3, ...)? ")) - 1
        which = file.sheet_names[index]


        while True:
            addmat = int(input("Would you like a material to a project (0) or display the current composition of a project(1)?"))
            if addmat == 0:
                addition()
            elif addmat == 1:
                plotmass(which)
            else:
                break
        break


    elif create == 1:
        firstcreation()
        break
    else:
        print("Please input either 0 or 1.")
        continue
while True:
    addmat = int(input("Would you like a material to a project (0) or display the current composition of a project(1)?"))
    if addmat == 0:
        addition()
    elif addmat == 1:
        project = registrar[which]
        plotmass(project)
    else:
        break